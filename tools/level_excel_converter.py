"""Convert PixelFarm level JSON files to/from an Excel workbook the art
team can edit without touching JSON or Godot.

Usage:
    python level_excel_converter.py <input> <output>

Direction is inferred from extensions:
    level_01.json -> level_01.xlsx   (JSON to Excel)
    level_01.xlsx -> level_01.json   (Excel to JSON)
"""

import json
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

TILE_TYPES = {"grass", "dirt", "water", "path"}
TILE_FILL_COLORS = {
    "grass": "408C33",
    "dirt": "8C5926",
    "water": "2666BF",
    "path": "B89E6B",
}
STATIC_PROP_TYPES = {"house", "sign"}

HEADER_FONT = Font(bold=True)


# ---- coordinate helpers ---------------------------------------------------

def pixel_to_grid(x, y, tile_size, width, height):
    return x / tile_size + width / 2, y / tile_size + height / 2


def grid_to_pixel(col, row, tile_size, width, height):
    x = round((col - width / 2) * tile_size)
    y = round((row - height / 2) * tile_size)
    return int(x), int(y)


def hex_to_rgb_floats(hex_color):
    hex_color = str(hex_color).lstrip("#")
    r = int(hex_color[0:2], 16) / 255
    g = int(hex_color[2:4], 16) / 255
    b = int(hex_color[4:6], 16) / 255
    return [round(r, 4), round(g, 4), round(b, 4)]


def rgb_floats_to_hex(rgb):
    r, g, b = (round(c * 255) for c in rgb[:3])
    return f"{r:02X}{g:02X}{b:02X}"


# ---- JSON -> Excel ---------------------------------------------------------

def json_to_excel(json_path: Path, xlsx_path: Path) -> None:
    data = json.loads(json_path.read_text(encoding="utf-8"))
    map_data = data.get("map", {})
    tile_size = map_data.get("tile_size", 64)
    width = map_data.get("width", 32)
    height = map_data.get("height", 24)

    wb = Workbook()
    wb.remove(wb.active)

    _write_meta_sheet(wb, data, map_data, tile_size, width, height)
    _write_tiles_sheet(wb, map_data.get("tiles", []), width, height)
    _write_spawns_sheet(wb, data.get("spawns", []), tile_size, width, height)
    _write_exits_sheet(wb, data.get("exits", []), tile_size, width, height)
    _write_props_sheets(wb, data.get("props", []), tile_size, width, height)

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)


def _write_meta_sheet(wb, data, map_data, tile_size, width, height):
    ws = wb.create_sheet("Meta")
    ws.append(["key", "value"])
    for cell in ws[1]:
        cell.font = HEADER_FONT
    for key, value in [
        ("id", data.get("id", "")),
        ("display_name", data.get("display_name", "")),
        ("music", data.get("music", "")),
        ("tile_size", tile_size),
        ("width", width),
        ("height", height),
        ("tileset", map_data.get("tileset", "")),
    ]:
        ws.append([key, value])


def _write_tiles_sheet(wb, zones, width, height):
    ws = wb.create_sheet("Tiles")
    half_w, half_h = width // 2, height // 2
    grid = [["" for _ in range(width)] for _ in range(height)]
    for zone in zones:
        zone_type = zone.get("type", "grass")
        for zx in range(zone["x"], zone["x"] + zone["w"]):
            for zy in range(zone["y"], zone["y"] + zone["h"]):
                col, row = zx + half_w, zy + half_h
                if 0 <= row < height and 0 <= col < width:
                    grid[row][col] = zone_type

    for row_index, row_values in enumerate(grid, start=1):
        for col_index, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.fill = PatternFill("solid", fgColor=TILE_FILL_COLORS.get(value or "grass"))


def _write_spawns_sheet(wb, spawns, tile_size, width, height):
    ws = wb.create_sheet("Spawns")
    ws.append(["id", "col", "row"])
    for cell in ws[1]:
        cell.font = HEADER_FONT
    for spawn in spawns:
        col, row = pixel_to_grid(spawn["x"], spawn["y"], tile_size, width, height)
        ws.append([spawn.get("id", ""), col, row])


def _write_exits_sheet(wb, exits, tile_size, width, height):
    ws = wb.create_sheet("Exits")
    ws.append(["col", "row", "target_level", "destination_spawn_id", "trigger"])
    for cell in ws[1]:
        cell.font = HEADER_FONT
    for exit_data in exits:
        col, row = pixel_to_grid(exit_data["x"], exit_data["y"], tile_size, width, height)
        ws.append([
            col, row,
            exit_data.get("target_level", ""),
            exit_data.get("destination_spawn_id", ""),
            exit_data.get("trigger", "interact"),
        ])


def _write_props_sheets(wb, props, tile_size, width, height):
    static_ws = wb.create_sheet("Props_Static")
    static_ws.append([
        "type", "col", "row", "texture", "color_hex",
        "collision_w", "collision_h", "collision_offset_y",
        "interact_prompt", "interact_action", "interact_target_level", "interact_text",
    ])
    item_ws = wb.create_sheet("Props_Item")
    item_ws.append(["col", "row", "item_name", "item_count", "color_hex"])
    tree_ws = wb.create_sheet("Props_Tree")
    tree_ws.append(["col", "row", "color_hex", "w", "h"])
    npc_ws = wb.create_sheet("Props_NPC")
    npc_ws.append(["col", "row", "npc_id"])

    for ws in (static_ws, item_ws, tree_ws, npc_ws):
        for cell in ws[1]:
            cell.font = HEADER_FONT

    for prop in props:
        prop_type = prop.get("type", "")
        col, row = pixel_to_grid(prop["x"], prop["y"], tile_size, width, height)

        if prop_type in STATIC_PROP_TYPES:
            interact = prop.get("interact", {})
            collision = prop.get("collision", {})
            color = prop.get("color")
            static_ws.append([
                prop_type, col, row,
                prop.get("texture") or "",
                rgb_floats_to_hex(color) if color else "",
                collision.get("w", ""),
                collision.get("h", ""),
                collision.get("offset_y", ""),
                interact.get("prompt", ""),
                interact.get("action", ""),
                interact.get("target_level", ""),
                interact.get("text", ""),
            ])
        elif prop_type == "item":
            color = prop.get("color")
            item_ws.append([
                col, row,
                prop.get("item_name", ""),
                prop.get("item_count", 1),
                rgb_floats_to_hex(color) if color else "",
            ])
        elif prop_type == "tree":
            color = prop.get("color")
            tree_ws.append([
                col, row,
                rgb_floats_to_hex(color) if color else "",
                prop.get("w", 16),
                prop.get("h", 16),
            ])
        elif prop_type == "npc":
            npc_ws.append([col, row, prop.get("npc_id", "")])
        else:
            raise ValueError(f"Unsupported prop type for Excel export: {prop_type!r}")


# ---- Excel -> JSON ---------------------------------------------------------

def excel_to_json(xlsx_path: Path, json_path: Path) -> None:
    wb = load_workbook(xlsx_path, data_only=True)

    meta = _read_meta_sheet(wb)
    tile_size, width, height = meta["tile_size"], meta["width"], meta["height"]

    data = {
        "id": meta["id"],
        "display_name": meta["display_name"],
        "music": meta["music"],
        "map": {
            "width": width,
            "height": height,
            "tileset": meta["tileset"],
            "tile_size": tile_size,
            "tiles": _read_tiles_sheet(wb, width, height),
        },
        "spawns": _read_spawns_sheet(wb, tile_size, width, height),
        "props": _read_props_sheets(wb, tile_size, width, height),
        "exits": _read_exits_sheet(wb, tile_size, width, height),
    }

    json_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.write_text(json.dumps(data, indent=2) + "\n", encoding="utf-8")


def _read_meta_sheet(wb):
    ws = wb["Meta"]
    values = {row[0].value: row[1].value for row in ws.iter_rows(min_row=2) if row[0].value}
    return {
        "id": values.get("id", ""),
        "display_name": values.get("display_name", ""),
        "music": values.get("music") or "",
        "tile_size": int(values.get("tile_size", 64)),
        "width": int(values.get("width", 32)),
        "height": int(values.get("height", 24)),
        "tileset": values.get("tileset") or "",
    }


def _read_tiles_sheet(wb, width, height):
    ws = wb["Tiles"]
    half_w, half_h = width // 2, height // 2
    zones = []
    for row_index in range(1, height + 1):
        for col_index in range(1, width + 1):
            cell = ws.cell(row=row_index, column=col_index)
            tile_type = str(cell.value or "grass").strip().lower()
            if tile_type == "grass":
                continue
            if tile_type not in TILE_TYPES:
                raise ValueError(f"Tiles!{cell.coordinate}: unknown tile type {tile_type!r}")
            zones.append({
                "x": (col_index - 1) - half_w,
                "y": (row_index - 1) - half_h,
                "w": 1,
                "h": 1,
                "type": tile_type,
            })
    return zones


def _read_spawns_sheet(wb, tile_size, width, height):
    ws = wb["Spawns"]
    spawns = []
    for row in ws.iter_rows(min_row=2):
        spawn_id = row[0].value
        if not spawn_id:
            continue
        x, y = grid_to_pixel(row[1].value, row[2].value, tile_size, width, height)
        spawns.append({"id": spawn_id, "x": x, "y": y})
    return spawns


def _read_exits_sheet(wb, tile_size, width, height):
    ws = wb["Exits"]
    exits = []
    for row in ws.iter_rows(min_row=2):
        target_level = row[2].value
        if not target_level:
            continue
        x, y = grid_to_pixel(row[0].value, row[1].value, tile_size, width, height)
        exits.append({
            "x": x,
            "y": y,
            "target_level": target_level,
            "destination_spawn_id": row[3].value or "",
            "trigger": row[4].value or "interact",
        })
    return exits


def _read_props_sheets(wb, tile_size, width, height):
    props = []

    static_ws = wb["Props_Static"]
    for row in static_ws.iter_rows(min_row=2):
        prop_type = row[0].value
        if not prop_type:
            continue
        x, y = grid_to_pixel(row[1].value, row[2].value, tile_size, width, height)
        prop = {"type": prop_type, "x": x, "y": y}
        texture, color_hex = row[3].value, row[4].value
        if texture:
            prop["texture"] = texture
        elif color_hex:
            prop["color"] = hex_to_rgb_floats(color_hex)
        collision = {}
        if row[5].value:
            collision["w"] = row[5].value
        if row[6].value:
            collision["h"] = row[6].value
        if row[7].value not in (None, ""):
            collision["offset_y"] = row[7].value
        if collision:
            prop["collision"] = collision
        interact = {}
        if row[8].value:
            interact["prompt"] = row[8].value
        if row[9].value:
            interact["action"] = row[9].value
        if row[10].value:
            interact["target_level"] = row[10].value
        if row[11].value:
            interact["text"] = row[11].value
        if interact:
            prop["interact"] = interact
        props.append(prop)

    item_ws = wb["Props_Item"]
    for row in item_ws.iter_rows(min_row=2):
        if row[0].value is None or row[1].value is None:
            continue
        x, y = grid_to_pixel(row[0].value, row[1].value, tile_size, width, height)
        prop = {
            "type": "item", "x": x, "y": y,
            "item_name": row[2].value or "Item",
            "item_count": row[3].value or 1,
        }
        if row[4].value:
            prop["color"] = hex_to_rgb_floats(row[4].value)
        props.append(prop)

    tree_ws = wb["Props_Tree"]
    for row in tree_ws.iter_rows(min_row=2):
        if row[0].value is None or row[1].value is None:
            continue
        x, y = grid_to_pixel(row[0].value, row[1].value, tile_size, width, height)
        prop = {"type": "tree", "x": x, "y": y}
        if row[2].value:
            prop["color"] = hex_to_rgb_floats(row[2].value)
        prop["w"] = row[3].value or 16
        prop["h"] = row[4].value or 16
        props.append(prop)

    npc_ws = wb["Props_NPC"]
    for row in npc_ws.iter_rows(min_row=2):
        if row[0].value is None or row[1].value is None:
            continue
        x, y = grid_to_pixel(row[0].value, row[1].value, tile_size, width, height)
        props.append({"type": "npc", "x": x, "y": y, "npc_id": row[2].value or "aiko"})

    return props


# ---- entry point ------------------------------------------------------------

def main():
    if len(sys.argv) != 3:
        print("Usage: python level_excel_converter.py <input> <output>")
        sys.exit(1)

    src, dst = Path(sys.argv[1]), Path(sys.argv[2])
    if src.suffix == ".json" and dst.suffix == ".xlsx":
        json_to_excel(src, dst)
    elif src.suffix == ".xlsx" and dst.suffix == ".json":
        excel_to_json(src, dst)
    else:
        print("Input/output must be .json -> .xlsx or .xlsx -> .json")
        sys.exit(1)

    print(f"Wrote {dst}")


if __name__ == "__main__":
    main()